#!/usr/bin/env python
# -*- encoding: UTF-8 -*-

"""
@File   : excel_tool.py
@Time   : 2020/11/25 8:57
@Author : yd
@Version: 1.0
@ToDo    : 使用openpyxl进行excel的读写操作 row column 从1开始
"""
import os
import shutil
import openpyxl
from openpyxl.styles import Font

from common import type_judgment
from common.logger import logger
from global_variables import get_abspath, cell_config, ui_cell_config


class ExcelTool:

    def __init__(self, file_path):
        self.file_path = get_abspath(file_path)
        if not os.path.basename(file_path).startswith("result_"):
            self.result_file_path = os.path.dirname(file_path) + os.sep + 'result_' + os.path.basename(file_path)
            shutil.copy(self.file_path, self.result_file_path)
        else:
            self.result_file_path = self.file_path

        self.workbook = openpyxl.load_workbook(self.result_file_path)
        self.sheet = None
        self.row = None
        self.column = None
        self.reading_row = 1
        self.sheet_name = None

    def get_sheet_names(self):
        """
        获取excel的sheet页
        :return: list
        """
        return self.workbook.sheetnames

    def set_sheet(self, sheet_name):
        """
        进行excel读写之前，需要设置要读取或者写入的sheet
        :param sheet_name:
        :return:
        """
        self.sheet = self.workbook[sheet_name]
        self.row = self.sheet.max_row
        self.column = self.sheet.max_column
        self.reading_row = 1

    def read_all(self):
        li = []
        for sheet_name in self.get_sheet_names():
            self.sheet_name = sheet_name
            self.set_sheet(sheet_name)
            list_1 = self.read_sheet()
            for l in list_1:
                li.append(l)
        return li

    def read_sheet(self):
        """
        获取当前sheet页所有行的
        :return:
        """
        lines = []
        # 遍历每行数据，跳过第一行的标题拦
        for i in range(2, self.row + 1):
            # line = {}
            line = []
            # if self.sheet.cell(i, cell_config.get('is_run')).value == '是':
            for j in range(1, self.column + 1):
                if self.sheet.cell(i, j).value is None:
                    line.append('')
                else:
                    if j == cell_config.get('method'):
                        line.append(str(self.sheet.cell(i, j).value).lower().replace('\n', ''))
                    else:
                        line.append(str(self.sheet.cell(i, j).value).replace('\n', ''))
            else:
                # 当取得所有值是，把sheet页写入list中，方便进行写操作时知道要写入那个sheet页
                line.append(self.sheet_name)
                # 把row加入lines中， 方便其他地方调用
                line.append(i)
            # print(line)
            lines.append(line)
        return lines

    def read_no_pass_line(self, row):
        line = []
        for j in range(1, self.column + 1):
            if cell_config.get('method') is not None or cell_config.get('method') == '':
                if self.sheet.cell(row, j).value is None:
                    line.append('')
                else:
                    line.append(str(self.sheet.cell(row, j).value).replace('\n', ''))
        return line

    def write(self, sheet_name, row=1, column=1, value='', color='000000'):
        """
        按行列写入对应的excel对的cell,写完后记得保存
        :param sheet_name: 需要写的sheet页
        :param row: 行
        :param column: 列
        :param value: 值
        :param color: 字体颜色
        :return:
        """
        row = int(row)
        column = int(column)
        if row > 0 and column > 0:
            # self.set_sheet(sheet_name)
            self.sheet = self.workbook[sheet_name]
            try:
                cell = self.sheet.cell(row=row, column=column, value=value)
                font = Font(name='Arial',
                            size=11,
                            bold=False,
                            italic=False,
                            vertAlign=None,
                            underline='none',
                            strike=False,
                            color=color)
                cell.font = font
            except Exception as e:
                logger.exception(e)
                self.sheet.cell(row=row, column=column, value=e)
            # self.save()
        else:
            logger.error('输入的xlsx的行和列必须大于1')

    def save(self):
        """
        写Excel后需要保存
        :return:
        """
        self.workbook.save(self.result_file_path)

    def read_ui_excel(self):
        cases = []
        step_list = []
        for sheet in self.get_sheet_names():
            self.set_sheet(sheet)
            li = []
            case = []
            for i in range(2, self.row + 1):
                if i == 2:
                    li.append(self.read_cell(i, ui_cell_config.get('id')))
                    li.append(self.read_cell(i, ui_cell_config.get('group')))
                    li.append(self.read_cell(i, ui_cell_config.get('case_name')))
                    li.append(self.read_cell(i, ui_cell_config.get('case_describe')))

                elif not type_judgment.is_Null(self.read_cell(i, ui_cell_config.get('id'))) \
                        and type_judgment.is_Null(self.read_cell(i, ui_cell_config.get('method'))):
                    if len(case) > 0:
                        li.append(case)
                        cases.append(li)
                        li = []
                        case = []

                    li.append(self.read_cell(i, ui_cell_config.get('id')))
                    li.append(self.read_cell(i, ui_cell_config.get('group')))
                    li.append(self.read_cell(i, ui_cell_config.get('case_name')))
                    li.append(self.read_cell(i, ui_cell_config.get('case_describe')))

                elif not type_judgment.is_Null(self.read_cell(i, ui_cell_config.get('method'))):
                    is_run = self.read_cell(i, ui_cell_config.get('is_run')).lower()
                    if is_run == '是' or is_run == 'yes' or is_run == 'y':
                        step_list = []
                        step_list.append(self.read_cell(i, ui_cell_config.get('case_step')))
                        step_list.append(self.read_cell(i, ui_cell_config.get('method')))
                        step_list.append(self.read_cell(i, ui_cell_config.get('expect_param1')))
                        step_list.append(self.read_cell(i, ui_cell_config.get('expect_param2')))
                        step_list.append(self.read_cell(i, ui_cell_config.get('expect_param3')))
                        step_list.append(self.read_cell(i, ui_cell_config.get('describe')))
                        step_list.append(i)
                        step_list.append(sheet)
                        # print('step =', step_list)
                        case.append(step_list)
                        # print('case =', case)
            else:
                if len(case) > 0:
                    li.append(case)
                    cases.append(li)
                    li = []
                    case = []
                    step_list = []
        return cases

    def read_cell(self, row, column):
        """
        调用前需要调用 set_sheet()方法
        :param row:
        :param column:
        :return:
        """
        row = int(row)
        column = int(column)
        value = self.sheet.cell(row, column).value
        if value is None:
            value = ''
        else:
            value = str(value).replace('\n', '').replace(' ', '').replace('\\', '')
        return value


if __name__ == '__main__':
    e = ExcelTool(get_abspath('data/cases/UI自动化测试用例.xlsx'))
    li = e.read_ui_excel()
    # print(li)
    for a in li:
        # print(a)
        lis = a[-1]
        for b in lis:
            print(b)
